#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
方案B 步骤①：将村委会导入 md_administrative_division (level 5)
方案B 步骤②：补全 md_residential_area.community_div_id

流程:
1. 从电子表格五级数据中筛选"村委会"行
2. 通过 code 匹配 level 4 (乡镇) 的 id 作为 parent_id，继承 path
3. 批量插入 md_administrative_division
4. UPDATE md_residential_area SET community_div_id = division.id WHERE code 匹配
"""

import os
import re
import glob
import openpyxl
import pymysql

# ============================================================
# 配置
# ============================================================
DATA_DIR = "/mnt/d/wsl/data/行政区划"
DB_HOST = "127.0.0.1"
DB_PORT = 3306
DB_USER = "root"
DB_PASS = "123456"
DB_NAME = "masterdata_db"
BATCH_SIZE = 1000
DRY_RUN = False


def load_division_map(conn):
    """加载行政区划 code → {id, parent_id, level, name, path}"""
    cursor = conn.cursor(pymysql.cursors.DictCursor)
    cursor.execute("""
        SELECT id, code, parent_id, level, name, path
        FROM md_administrative_division
        WHERE delete_time IS NULL
    """)
    result = {}
    for row in cursor.fetchall():
        result[row['code']] = row
    cursor.close()
    return result


def parse_xlsx(filepath):
    """解析单个 xlsx 文件的五级数据，返回村委会行"""
    rows = []
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
    except Exception as e:
        print(f"  [WARN] 无法打开 {os.path.basename(filepath)}: {e}")
        return rows

    if '五级数据' not in wb.sheetnames:
        wb.close()
        return rows

    ws = wb['五级数据']
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0] or not row[8]:
            continue
        name = str(row[10] or '').strip()
        if '村委会' not in name:
            continue
        rows.append({
            'street_code': str(row[6] or '').strip() if row[6] else None,
            'village_code': str(row[8] or '').strip(),
            'name': name,
            'urban_rural': str(row[9] or '').strip() if row[9] else None,
        })

    wb.close()
    return rows


def main():
    xlsx_files = sorted(glob.glob(os.path.join(DATA_DIR, "*.xlsx")))
    xlsx_files = [f for f in xlsx_files if not os.path.basename(f).startswith('~$')]
    print(f"找到 {len(xlsx_files)} 个 xlsx 文件")

    conn = pymysql.connect(
        host=DB_HOST, port=DB_PORT, user=DB_USER, password=DB_PASS,
        database=DB_NAME, charset='utf8mb4', autocommit=False
    )

    print("加载行政区划映射...")
    div_map = load_division_map(conn)
    print(f"  行政区划记录数: {len(div_map)}")

    # ---- 步骤1: 收集村委会数据 ----
    total_villages = 0
    skipped_no_street = 0
    skipped_exist = 0
    insert_list = []
    seen_codes = set()

    # 预加载已有的 code（包括已经插入过的）
    for code, info in div_map.items():
        seen_codes.add(code)

    for filepath in xlsx_files:
        fname = os.path.basename(filepath)
        villages = parse_xlsx(filepath)
        if not villages:
            continue

        file_matched = 0
        for v in villages:
            total_villages += 1

            # 去重
            if v['village_code'] in seen_codes:
                skipped_exist += 1
                continue

            # 查找 parent (level 4 乡镇/街道)
            street = div_map.get(v['street_code'])
            if not street or street['level'] != 4:
                skipped_no_street += 1
                continue

            seen_codes.add(v['village_code'])
            file_matched += 1

            insert_list.append({
                'code': v['village_code'],
                'name': v['name'],
                'parent_id': street['id'],
                'path': f"{street['path']}",  # 临时，插入后再补
                'street_div_id': street['id'],
                'urban_rural': v['urban_rural'],
            })

        if file_matched > 0:
            print(f"  {fname}: +{file_matched} 条村委会")

    print(f"\n{'='*60}")
    print(f"统计:")
    print(f"  电子表格村委会总数: {total_villages}")
    print(f"  待插入:             {len(insert_list)}")
    print(f"  跳过-街道未找到:    {skipped_no_street}")
    print(f"  跳过-编码已存在:    {skipped_exist}")

    if DRY_RUN:
        print(f"\n[DRY RUN] 前5条:")
        for item in insert_list[:5]:
            parent = div_map.get(str(item['parent_id']), {})
            print(f"  {item['code']} | parent={item['parent_id']}({parent.get('name','?')}) | {item['name']}")
        conn.close()
        return

    if not insert_list:
        print("\n没有需要插入的村委会数据")
        conn.close()
        return

    # ---- 步骤1: 批量插入 md_administrative_division ----
    print(f"\n步骤1: 插入 {len(insert_list)} 条村委会到行政区划表...")

    sql_insert_div = """
    INSERT INTO md_administrative_division
        (parent_id, level, name, code, path, sort_order, status, urban_rural_code, created_by)
    VALUES
        (%s, 5, %s, %s, %s, 0, 1, %s, 1)
    """

    cursor = conn.cursor()
    inserted = 0
    # 需要在插入后获取 id 来构建 path，所以逐批处理
    code_to_ra = {}  # village_code → ra id (用于步骤2)

    for i in range(0, len(insert_list), BATCH_SIZE):
        batch = insert_list[i:i + BATCH_SIZE]
        params = [(
            item['parent_id'], item['name'], item['code'],
            item['path'], item['urban_rural'],
        ) for item in batch]

        try:
            cursor.executemany(sql_insert_div, params)
            conn.commit()
            inserted += len(batch)
            if inserted % 10000 == 0 or inserted == len(insert_list):
                print(f"  进度: {inserted}/{len(insert_list)}")
        except Exception as e:
            print(f"  [ERROR] 批次插入失败: {e}")
            conn.rollback()
            cursor.close()
            conn.close()
            return

    print(f"  ✅ 行政区划表插入完成: {inserted} 条")

    # ---- 步骤1.5: 修正 path（需要拿到自增 id） ----
    print(f"\n修正 path...")
    cursor.execute("""
        SELECT id, code FROM md_administrative_division
        WHERE level = 5 AND name LIKE '%村委会' AND delete_time IS NULL
    """)
    div_rows = cursor.fetchall()

    if div_rows:
        # 构建 code → 新插入的 division id
        new_div_ids = {row[1]: row[0] for row in div_rows}

        update_count = 0
        for item in insert_list:
            new_id = new_div_ids.get(item['code'])
            if not new_id:
                continue
            parent = div_map.get(str(item['parent_id']), {})
            parent_path = parent.get('path', '/')
            new_path = f"{parent_path}{new_id}/"
            cursor.execute(
                "UPDATE md_administrative_division SET path = %s WHERE id = %s",
                (new_path, new_id)
            )
            update_count += 1

        conn.commit()
        print(f"  ✅ 修正 {update_count} 条 path")

    # ---- 步骤2: 补全 md_residential_area.community_div_id ----
    print(f"\n步骤2: 补全 md_residential_area.community_div_id...")

    cursor.execute("""
        SELECT id, code FROM md_residential_area
        WHERE community_type = 2 AND (community_div_id IS NULL OR community_div_id = 0)
        AND delete_time IS NULL
    """)
    ra_rows = cursor.fetchall()

    # 加载 division code→id (包括刚插入的)
    cursor.execute("SELECT code, id FROM md_administrative_division WHERE level=5 AND delete_time IS NULL")
    div_code_to_id = {row[0]: row[1] for row in cursor.fetchall()}

    matched_count = 0
    unmatched_count = 0
    batch_params = []

    for ra_id, ra_code in ra_rows:
        div_id = div_code_to_id.get(ra_code)
        if div_id:
            batch_params.append((div_id, ra_id))
            matched_count += 1
        else:
            unmatched_count += 1

    if batch_params:
        for i in range(0, len(batch_params), BATCH_SIZE):
            b = batch_params[i:i + BATCH_SIZE]
            cursor.executemany(
                "UPDATE md_residential_area SET community_div_id = %s WHERE id = %s",
                b
            )
        conn.commit()

    print(f"  匹配成功: {matched_count}")
    print(f"  未匹配:   {unmatched_count}")
    print(f"  ✅ community_div_id 补全完成")

    cursor.close()
    conn.close()
    print(f"\n🎉 全部完成!")


if __name__ == '__main__':
    main()
