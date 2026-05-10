#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
导入村委会数据到 md_residential_area 表
- 从行政区划电子表格中筛选"村委会"行
- 通过 code 匹配 md_administrative_division 获取 county_id / street_id
- 村级编码作为 md_residential_area.code
"""

import os
import re
import glob
import openpyxl
import pymysql

# ============================================================
# 配置
# ============================================================
DATA_DIR = "/mnt/d/wsl/data/行政区划"
DB_HOST = "127.0.0.1"
DB_PORT = 3306
DB_USER = "root"
DB_PASS = "123456"
DB_NAME = "masterdata_db"
BATCH_SIZE = 1000
DRY_RUN = False  # 设为 True 只打印统计不执行插入

# ============================================================
# 读取行政区划表的 code→id 映射（一次性加载）
# ============================================================
def load_division_map(conn):
    """返回 {code: id, level, name} 的字典"""
    cursor = conn.cursor(pymysql.cursors.DictCursor)
    cursor.execute("""
        SELECT id, code, level, name
        FROM md_administrative_division
        WHERE delete_time IS NULL
    """)
    result = {}
    for row in cursor.fetchall():
        result[row['code']] = row
    cursor.close()
    return result


def load_existing_ra_codes(conn):
    """返回已存在的 residential_area code 集合"""
    cursor = conn.cursor()
    cursor.execute("SELECT code FROM md_residential_area WHERE code IS NOT NULL AND delete_time IS NULL")
    result = {row[0] for row in cursor.fetchall()}
    cursor.close()
    return result


# ============================================================
# 解析电子表格
# ============================================================
def parse_xlsx(filepath):
    """解析单个 xlsx 文件的五级数据 sheet，返回村委会行列表"""
    rows = []
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
    except Exception as e:
        print(f"  [WARN] 无法打开 {os.path.basename(filepath)}: {e}")
        return rows

    if '五级数据' not in wb.sheetnames:
        wb.close()
        return rows

    ws = wb['五级数据']

    for row in ws.iter_rows(min_row=3, values_only=True):
        # 跳过空行
        if not row[0] or not row[8]:
            continue

        name = str(row[10] or '').strip()
        # 只筛选"村委会"
        if '村委会' not in name:
            continue

        rows.append({
            'province_code': int(row[0]) if row[0] else None,
            'city_code': int(row[2]) if row[2] else None,
            'county_code': int(row[4]) if row[4] else None,
            'county_name': str(row[5] or '').strip(),
            'street_code': int(row[6]) if row[6] else None,
            'street_name': str(row[7] or '').strip(),
            'village_code': str(row[8] or '').strip(),
            'urban_rural': str(row[9] or '').strip(),
            'name': name,
        })

    wb.close()
    return rows


# ============================================================
# 主流程
# ============================================================
def main():
    # 扫描所有 xlsx 文件
    xlsx_files = sorted(glob.glob(os.path.join(DATA_DIR, "*.xlsx")))
    xlsx_files = [f for f in xlsx_files if os.path.basename(f).startswith('~$') is False]
    print(f"找到 {len(xlsx_files)} 个 xlsx 文件")

    # 连接数据库
    conn = pymysql.connect(
        host=DB_HOST, port=DB_PORT, user=DB_USER, password=DB_PASS,
        database=DB_NAME, charset='utf8mb4', autocommit=False
    )

    print("加载行政区划映射...")
    div_map = load_division_map(conn)
    print(f"  行政区划记录数: {len(div_map)}")

    print("加载已有小区编码...")
    existing_codes = load_existing_ra_codes(conn)
    print(f"  已有小区编码数: {len(existing_codes)}")

    # 统计
    total_villages = 0
    matched = 0
    skipped_exist = 0
    skipped_no_county = 0
    skipped_no_street = 0
    insert_list = []

    for filepath in xlsx_files:
        fname = os.path.basename(filepath)
        villages = parse_xlsx(filepath)
        if not villages:
            continue

        print(f"\n  处理 {fname}: {len(villages)} 条村委会数据")

        for v in villages:
            total_villages += 1
            county_code = str(v['county_code'])
            street_code = str(v['street_code'])

            # 查找区县 ID
            county_div = div_map.get(county_code)
            if not county_div or county_div['level'] != 3:
                skipped_no_county += 1
                continue

            # 查找街道 ID
            street_div = div_map.get(street_code)
            if not street_div or street_div['level'] != 4:
                skipped_no_street += 1
                continue

            # 去重：已存在则跳过
            if v['village_code'] in existing_codes:
                skipped_exist += 1
                continue

            # 构建名称：去掉"村委会"后缀
            clean_name = re.sub(r'村委会$', '', v['name']).strip()

            matched += 1

            # 内存去重（避免跨文件重复，如北京市有2个相同文件）
            if v['village_code'] in existing_codes:
                skipped_exist += 1
                continue
            existing_codes.add(v['village_code'])

            insert_list.append({
                'code': v['village_code'],
                'name': clean_name,
                'county_id': county_div['id'],
                'street_id': street_div['id'],
                'community_div_id': None,  # 村委会通常无社区层
                'address': f"{v['county_name']}{v['street_name']}{clean_name}",
                'area': None,
                'population': None,
                'community_type': 2,  # 2=村庄
                'submission_status': 2,  # 2=已审批（导入数据直接通过）
                'submitter_id': 1,
            })

    print(f"\n{'='*60}")
    print(f"统计:")
    print(f"  电子表格中村委会总数: {total_villages}")
    print(f"  匹配成功(区县+街道):  {matched}")
    print(f"  跳过-区县未找到:      {skipped_no_county}")
    print(f"  跳过-街道未找到:      {skipped_no_street}")
    print(f"  跳过-编码已存在:      {skipped_exist}")

    if DRY_RUN:
        print(f"\n[DRY RUN] 不执行插入，显示前10条:")
        for item in insert_list[:10]:
            print(f"  {item['code']} | county_id={item['county_id']} | street_id={item['street_id']} | {item['name']}")
        conn.close()
        return

    if not insert_list:
        print("\n没有需要插入的数据")
        conn.close()
        return

    # 批量插入
    print(f"\n开始批量插入 {len(insert_list)} 条数据...")

    sql = """
    INSERT INTO md_residential_area
        (code, name, county_id, street_id, community_div_id, address,
         area, population, community_type, submission_status, submitter_id,
         created_time, updated_time)
    VALUES
        (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
    """

    cursor = conn.cursor()
    inserted = 0
    for i in range(0, len(insert_list), BATCH_SIZE):
        batch = insert_list[i:i + BATCH_SIZE]
        params = [(
            item['code'], item['name'], item['county_id'], item['street_id'],
            item['community_div_id'], item['address'], item['area'], item['population'],
            item['community_type'], item['submission_status'], item['submitter_id'],
        ) for item in batch]
        try:
            cursor.executemany(sql, params)
            inserted += len(batch)
            print(f"  进度: {inserted}/{len(insert_list)}")
        except Exception as e:
            print(f"  [ERROR] 批次 {i//BATCH_SIZE+1} 插入失败: {e}")
            conn.rollback()
            cursor.close()
            conn.close()
            return

    conn.commit()
    cursor.close()
    conn.close()
    print(f"\n✅ 完成! 共插入 {inserted} 条村委会数据")


if __name__ == '__main__':
    main()
